import subprocess
import sys
import os

def install_and_import(package, import_as=None):
    if not import_as:
        import_as = package
    try:
        __import__(import_as)
        print(f"{import_as} is already installed.")
    except ImportError:
        print(f"{import_as} is not installed. Attempting installation...")
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", package])
            print(f"{package} was successfully installed.")
            globals()[import_as] = __import__(import_as)
        except Exception as e:
            print(f"Error during the installation of {package}: {e}")

libraries = {
    're': 're',  
    'yfinance': 'yfinance', 
    'matplotlib': 'matplotlib.pyplot',  
    'datetime': 'datetime',
    'openpyxl': 'openpyxl',
    'uuid' : 'uuid'
}

for package, import_as in libraries.items():
    install_and_import(package, import_as)

import matplotlib.pyplot as plt
import datetime
import yfinance
import uuid

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image

from datetime import date

print()

#####################################################################################################################
class UserAccount:
    existing_usernames = set()
    
    excel_file = 'user_accounts.xlsx'
    # Encapsulation : attribut password can only be printed by the method get_password because it's a private attribute (__), data is hide from the outside for data integrity
    def __init__(self, username, password, email, first_name=None, last_name=None):
        if username in UserAccount.existing_usernames:
            raise ValueError(f"Username '{username}' is already taken. Please choose a different username.")
        self.username = username
        UserAccount.existing_usernames.add(username)
        self.__password = password  
        self.email = email
        self.first_name = first_name
        self.last_name = last_name
        self.portfolios = []
        self.save_user_to_excel()
        
    excel_file = 'CryptoInvestmentData.xlsx'
    sheet_name = 'User Info'
    
    def get_portfolios_string(self):
        return ', '.join(portfolio.name for portfolio in self.portfolios)

    def save_user_to_excel(self):
        try:
            wb = openpyxl.load_workbook(UserAccount.excel_file)
            if UserAccount.sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(UserAccount.sheet_name)
                ws.append(['Username', 'Email', 'First Name', 'Last Name', 'Total Liquidity', 'Portfolios', 'Portfolio IDs'])
            else:
                ws = wb[UserAccount.sheet_name]
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            ws.title = UserAccount.sheet_name

            ws.append(['Username', 'Email', 'First Name', 'Last Name', 'Total Liquidity', 'Portfolios', 'Portfolio IDs'])
        
        existing_row_index = None
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] == self.username:
                existing_row_index = idx
                break

        portfolio_ids = ', '.join(str(portfolio.id) for portfolio in self.portfolios)
        if existing_row_index:
            ws.cell(row=existing_row_index, column=2, value=self.email)
            ws.cell(row=existing_row_index, column=3, value=self.first_name if self.first_name else '')
            ws.cell(row=existing_row_index, column=4, value=self.last_name if self.last_name else '')
            ws.cell(row=existing_row_index, column=5, value=self.view_total_liquidity())
            ws.cell(row=existing_row_index, column=6, value=self.get_portfolios_string())
            ws.cell(row=existing_row_index, column=7, value=portfolio_ids) 
        else:
            ws.append([
                self.username,
                self.email,
                self.first_name if self.first_name else '',
                self.last_name if self.last_name else '',
                self.view_total_liquidity(),
                self.get_portfolios_string(),
                portfolio_ids  
            ])

        wb.save(UserAccount.excel_file)
    
    def view_total_liquidity(self):
        return sum(portfolio.liquidity for portfolio in self.portfolios)

    def set_username(self, new_username):
        if new_username in UserAccount.existing_usernames:
            print(f"Username '{new_username}' is already taken. Please choose a different username.")
            return False

        UserAccount.existing_usernames.discard(self.username)
        self.username = new_username
        UserAccount.existing_usernames.add(new_username)
        return True
    
    def set_password(self, new_password):
        self.__password = new_password
    
    def set_email(self, new_email):
        self.email = new_email
    
    def set_first_name(self, new_first_name):
        self.first_name = new_first_name
    
    def set_last_name(self, new_last_name):
        self.last_name = new_last_name
    
    def add_portfolio(self, new_portfolio):
        if any(p.name == new_portfolio.name for p in self.portfolios):
            print(f"A portfolio named '{new_portfolio.name}' already exists for user {self.username}.")
            return False
        self.portfolios.append(new_portfolio)
        print(f"Portfolio '{new_portfolio.name}' added successfully for user {self.username}.")
        return True

    def remove_portfolio(self, portfolio_to_remove):
        for idx, portfolio in enumerate(self.portfolios):
            if portfolio.name == portfolio_to_remove.name:
                del self.portfolios[idx] 
                print(f"The portfolio '{portfolio_to_remove.name}' was successfully removed.")
                return True
        print(f"No portfolio named '{portfolio_to_remove.name}' found.")
        return False
    
    def get_user_info(self):
        info = f"Username: {self.username}, Email: {self.email}, First Name: {self.first_name}, Last Name: {self.last_name}"
        if self.portfolios:
            info += "\nPortfolios:"
            for portfolio in self.portfolios:
                info += f"\n- {portfolio.name}"
            total_liquidity = self.view_total_liquidity()
            info += f"\nTotal liquidity available across all portfolios: {total_liquidity} USD"
        else:
            info += "\nNo portfolio created."
        
        print(info)
        return info

"""
user1 = UserAccount("johnDoe123", "supersecretpassword", "johndoe@example.com", "John", "Doe")
print(user1.get_user_info())

#Updating user information
user1.set_email("newemail@example.com")
user1.set_first_name("Johnny")
user1.set_last_name("Doe II")
print(user1.get_user_info())
user1.save_user_to_excel()
"""

################################################################################################################################################################
"""
ticker = yfinance.Ticker("BTC-USD")
info = ticker.info
Imprime toutes les clés et leurs valeurs associées
for key, value in info.items():
   print(f"{key}: {value}")
"""
class Crypto:
    def __init__(self, ticker):
        self.ticker = ticker
        self.info = None
        self.load_info()

    def load_info(self):
        try:
            self.info = yfinance.Ticker(self.ticker).info
        except Exception as e:
            print(f"Error loading information for {self.ticker}: {e}")
    
    # Abstraction : details of the extraction by the yfinance API is hidden, simplifies complexity
    def get_price_close(self):
        """Returns the current price."""
        return self.info.get('previousClose') if self.info else None

    def get_volume(self):
        """Returns the current volume."""
        return self.info.get('regularMarketVolume') if self.info else None
    
    def get_market_cap(self):
        """Returns the market cap."""
        return self.info.get('marketCap') if self.info else "Market cap information not available."
        
    def get_description(self):
        """Returns the description."""
        return self.info.get('description') if self.info else None

    def extract_price_from_description(self):
        """Extracts the price from the description."""
        description = self.get_description()
        if description:
            match = re.search(r"The last known price of Bitcoin is ([\d,]+.\d+) USD", description)
            if match:
                return match.group(1)
            else:
                return "Price not found in description."
        else:
            return "Description not available."

    def analyze_investment_opportunity(self, show_plot=False):
        today = datetime.date.today()
        today_str = today.strftime('%Y-%m-%d')
        data = yfinance.download(self.ticker, start="2020-01-01", end=today_str)
        data['MA120'] = data['Adj Close'].rolling(window=120).mean()
        data['MA240'] = data['Adj Close'].rolling(window=240).mean()

        if data['MA120'].iloc[-1] > data['MA240'].iloc[-1]:
            decision = "Invest Advice: Buy MA120 > MA240"
            decision_color = "green"
        else:
            decision = "Invest Advice: Sell MA120 < MA240"
            decision_color = "red"

        figure_path = f"Analysis/{self.ticker}.png"
        os.makedirs(os.path.dirname(figure_path), exist_ok=True)

        plt.figure(figsize=(10, 6))
        plt.plot(data.index, data['Adj Close'], label='Adj Close')
        plt.plot(data.index, data['MA120'], label='MA 120 Days', color='green', linestyle='--')
        plt.plot(data.index, data['MA240'], label='MA 240 Days', color='red', linestyle='--')
        plt.title(f'{self.ticker} Investment Analysis')
        plt.figtext(0.5, 0.01, decision, wrap=True, horizontalalignment='center', fontsize=12, color=decision_color)  # Ajouter le sous-titre ici
        plt.legend()

        plt.savefig(figure_path)

        if show_plot:
            plt.show()
        plt.close()

        return decision, figure_path
        
def save_crypto_data_to_excel(crypto_data, filename='CryptoInvestmentData.xlsx', sheet_name='Crypto Info'):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    ws.column_dimensions[get_column_letter(1)].width = 15
    ws.column_dimensions[get_column_letter(2)].width = 15
    ws.column_dimensions[get_column_letter(3)].width = 10
    ws.column_dimensions[get_column_letter(4)].width = 20
    ws.column_dimensions[get_column_letter(5)].width = 30
    ws.column_dimensions[get_column_letter(6)].width = 20
    ws.column_dimensions[get_column_letter(7)].width = 40
    ws.append(['Ticker', 'Previous Close', 'Volume', 'Market Cap', 'Description', 'Investment Decision', 'Chart Analysis'])

    for index, data in enumerate(crypto_data, start=1):
        ws.append([
            data['ticker'],
            data['previous_close'],
            data['volume'],
            data['market_cap'],
            data['description'],
            data['investment_decision']
        ])

        img_path = data['Analysis_image']
        if os.path.exists(img_path):
            img = Image(img_path)
            img.width, img.height = img.width * 0.25, img.height * 0.25
            cell_ref = f'G{ws.max_row}'
            ws.add_image(img, cell_ref)

    wb.save(filename)

# Polymorphism : mainly classes trated by the instance of the Crypto class through inheritance for the 10 most popular cryptoccurencies
tickers = [
    'BTC-USD', 'ETH-USD', 'BNB-USD', 'XRP-USD', 'SOL-USD',
    'ADA-USD', 'DOT-USD', 'DOGE-USD', 'LTC-USD', 'LINK-USD'
]

class CryptoFactory:
    @staticmethod # We use the class for this method not the object
    # Factory Method : Here we create classes instances without specifying the exact class of the object created with the CryptoFactory class, no constructor
    def create_crypto(ticker):
        return Crypto(ticker)

crypto_data_list = []
factory = CryptoFactory()

for ticker in tickers:
    # Utilization of the factory
    crypto = factory.create_crypto(ticker)
    decision, Analysis_image = crypto.analyze_investment_opportunity()
    crypto_data = {
        'ticker': ticker,
        'previous_close': crypto.get_price_close(),
        'volume': crypto.get_volume(),
        'market_cap': crypto.get_market_cap(),
        'description': crypto.get_description(),
        'investment_decision': decision,
        'Analysis_image': Analysis_image
    }
    crypto_data_list.append(crypto_data)


# Inheritance : subclass derivated from the Crypto parent class with same attributes (and methods witch can be overrides)
class Bitcoin(Crypto):
    def __init__(self):
        super().__init__('BTC-USD')

    """
    print(f"{ticker} Price Close: {crypto.get_price_close()} USD")
    print(f"{ticker} Volume: {crypto.get_volume()}")
    # print(f"{ticker} Description: {crypto.get_description()}")
    print(f"{ticker} Market Cap: {crypto.get_market_cap()} USD")
    print("\n" + "-"*50 + "\n")
    """

#######################################################################################################
class Portfolio:
    def __init__(self, name):
        self.id = str(uuid.uuid4())
        self.name = name
        self.liquidity = 0
        self.crypto_balances = {}
        self.transaction_history = []
        
    def save_portfolio_to_excel(self, filename='CryptoInvestmentData.xlsx'):
        try:
            wb = load_workbook(filename)
            if 'Portfolios' in wb.sheetnames:
                ws = wb['Portfolios']
                if ws['A1'].value != 'ID':
                    ws.insert_rows(1)
                    ws.append(['ID', 'Name', 'Liquidity', 'Crypto Balances', 'Transaction History'])
            else:
                ws = wb.create_sheet('Portfolios')
                ws.append(['ID', 'Name', 'Liquidity', 'Crypto Balances', 'Transaction History'])
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            ws.title = 'Portfolios'
            ws.append(['ID', 'Name', 'Liquidity', 'Crypto Balances', 'Transaction History'])

        id_exists = False
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if self.id == row[0]:
                id_exists = True
                break

        if not id_exists:
            ws.append([
                self.id, 
                self.name, 
                self.liquidity, 
                ', '.join(f"{ticker}: {quantity}" for ticker, quantity in self.crypto_balances.items()), 
                '; '.join(self.transaction_history)
            ])
        else:
            for row in ws.iter_rows(min_row=2):
                if row[0].value == self.id:
                    row[1].value = self.name
                    row[2].value = self.liquidity
                    row[3].value = ', '.join(f"{ticker}: {quantity}" for ticker, quantity in self.crypto_balances.items())
                    row[4].value = '; '.join(self.transaction_history)
                    break

        wb.save(filename)
        
    def set_name(self, new_name):
        self.name = new_name
        print(f"The portfolio's name has been changed to '{new_name}'.")

    def view_transaction_history(self):
        if self.transaction_history:
            print("Transaction history:")
            for transaction in self.transaction_history:
                print(transaction)
        else:
            print("No transactions have been made.")

    def get_portfolio_info(self):
        info = f"Portfolio Name: {self.name}\nAvailable Liquidity: {self.liquidity} USD\n"
        if self.crypto_balances:
            info += "Cryptocurrency Balances:\n"
            for ticker, quantity in self.crypto_balances.items():
                info += f"  {ticker}: {quantity}\n"
        else:
            info += "No cryptocurrencies in this portfolio.\n"
        info += "\nTransaction History:\n" + ("\n".join(self.transaction_history) if self.transaction_history else "No transactions made.")
        print(info)

    def add_liquidity(self):
        amount = input("How much would you like to add to the liquidity? ")
        try:
            amount = float(amount)
            if amount > 0:
                self.liquidity += amount
                self.transaction_history.append(f"Added {amount} USD to liquidity")
                print(f"{amount} USD added to liquidity. Total liquidity: {self.liquidity} USD")
            else:
                print("Please enter a positive amount.")
        except ValueError:
            print("Please enter a valid number.")

    def withdraw_liquidity(self):
        amount = input("How much would you like to withdraw from the liquidity? ")
        try:
            amount = float(amount)
            if amount > 0 and self.liquidity >= amount:
                self.liquidity -= amount
                self.transaction_history.append(f"Withdrew {amount} USD from liquidity")
                print(f"{amount} USD withdrawn from liquidity. Remaining liquidity: {self.liquidity} USD")
            elif amount <= 0:
                print("Please enter a positive amount.")
            else:
                print("Insufficient balance.")
        except ValueError:
            print("Please enter a valid number.")

    def view_balances(self):
        print(f"Available Liquidity: {self.liquidity} USD")
        print("Cryptocurrency Balances:")
        for ticker, quantity in self.crypto_balances.items():
            print(f"  {ticker}: {quantity}")

    def buy_crypto(self):
        available_tickers = ['BTC-USD', 'ETH-USD', 'BNB-USD', 'XRP-USD', 'SOL-USD', 'ADA-USD', 'DOT-USD', 'DOGE-USD', 'LTC-USD', 'LINK-USD']
        while True:
            ticker = input(f"Which crypto would you like to buy from {', '.join(available_tickers)}? Enter the ticker or 'cancel' to exit: ").strip()
            if ticker == 'cancel':
                print("Operation cancelled.")
                break
            if ticker not in available_tickers:
                print("This ticker is not in the list of available cryptos. Please try again or cancel.")
                continue

            try:
                total_amount = float(input("Enter the total amount in USD you wish to spend: "))
            except ValueError:
                print("Please enter a valid number.")
                continue

            crypto = Crypto(ticker)
            current_price = crypto.get_price_close()
            if current_price is None:
                print(f"Unable to retrieve the current price for {ticker}. Please try again.")
                continue

            fee = total_amount * 0.005 
            amount_after_fees = total_amount - fee

            if self.liquidity >= total_amount:
                purchasable_quantity = amount_after_fees / current_price
                self.crypto_balances[ticker] = self.crypto_balances.get(ticker, 0) + purchasable_quantity
                self.liquidity -= total_amount
                Platform.collect_fees(fee)
                print(f"Purchase successful: {purchasable_quantity} of {ticker} for {amount_after_fees} USD (0.05% fee included).")
                print(f"Remaining liquidity: {self.liquidity} USD")
                break
            else:
                print("Purchase failed. Insufficient liquid balance.")
                break
                
    def sell_crypto(self):
        if not self.crypto_balances:
            print("Your cryptocurrency portfolio is empty.")
            return

        while True:
            print("Cryptos available to sell in your portfolio:")
            for ticker, quantity in self.crypto_balances.items():
                print(f"{ticker}: {quantity}")

            ticker = input("Enter the ticker of the crypto you wish to sell or 'cancel' to exit: ").strip()

            if ticker.lower() == 'cancel':
                print("Operation cancelled.")
                break

            if ticker not in self.crypto_balances:
                print("This ticker is not present in your portfolio. Please try again.")
                continue

            quantity_to_sell = self.crypto_balances[ticker]

            crypto = Crypto(ticker)
            current_price = crypto.get_price_close()
            if current_price is None:
                print(f"Unable to retrieve the current price for {ticker}. Please try again.")
                continue

            amount_received = quantity_to_sell * current_price
            self.liquidity += amount_received
            del self.crypto_balances[ticker]
            print(f"Sale successful: You sold {quantity_to_sell} {ticker} for a total of {amount_received} USD")
            self.transaction_history.append(f"Sale of {quantity_to_sell} {ticker} received {amount_received} USD")
            print(f"Remaining liquidity: {self.liquidity} USD")
            break

#####################################################################################################################################
# Singleton : only one instance for the platform (_intsance, classmethod)
class Platform: 
    _instance = None
    total_fees = 0.0
    users = []

    def __new__(cls, name, siret, location):
        if cls._instance is None:
            cls._instance = super().__new__(cls)
            cls._instance.name = name
            cls._instance.siret = siret
            cls._instance.location = location
        return cls._instance

    @classmethod
    def collect_fees(cls, fees):
        cls.total_fees += fees

    def add_user(self, user):
        self.users.append(user)
        print(f"User {user} added to the platform.")

    def remove_user(self, user):
        self.users = [u for u in self.users if u != user]
        print(f"User {user} removed from the platform.")

    @classmethod
    def pay_taxes(cls):
        taxes = cls.total_fees * 0.30
        cls.total_fees -= taxes

    @classmethod
    def save_accounting_to_excel(self, filename='CryptoInvestmentData.xlsx'):
        today_str = date.today().isoformat()
        try:
            wb = load_workbook(filename)
            if 'Accounting' in wb.sheetnames:
                ws = wb['Accounting']
            else:
                ws = wb.create_sheet('Accounting')
                ws.append(['Type', 'Amount (USDT)', 'Description', 'Date'])
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            ws.title = 'Accounting'
            ws.append(['Type', 'Amount', 'Description', 'Date'])

        existing_entry = False
        for row in ws.iter_rows(min_row=2, max_col=4, values_only=True):
            if row[3] == today_str:
                existing_entry = True
                break

        if not existing_entry:
            fees_row = ['Fees', self.total_fees, 'Total fees accumulated', today_str]
            taxes = self.total_fees * 0.30
            taxes_row = ['Taxes', -taxes, 'Taxes paid', today_str]
            net_profit = self.total_fees - taxes
            net_profit_row = ['Net Profit', net_profit, 'Net profit after taxes', today_str]

            ws.append(fees_row)
            ws.append(taxes_row)
            ws.append(net_profit_row)

        wb.save(filename)
        if not existing_entry:
            print("Accounting data saved to Excel.")
        else:
            print("Today's accounting data already exists in Excel.")

#################################################################################################################################
# Test
#################################################################################################################################
platform = Platform("CryptoPlatform", 36252187900034, "Vilnius, Lithuania")
#################################################################################################################################    
user1 = UserAccount("romTaug20", "IloveVilniusTech$$", "romtaug@gmail.com", "Romain", "Taugourdeau")
platform.add_user(user1)
user1.get_user_info()
###########################################################################################################################
print()
wallet1 = Portfolio("wallet1")
wallet2 = Portfolio("wallet2")
user1.add_portfolio(wallet1)
user1.get_user_info()
print()
user1.add_portfolio(wallet2)
user1.get_user_info()
print()
user1.remove_portfolio(wallet2)
user1.get_user_info()
####################################################################################################################################
print()
wallet1.add_liquidity()
print()
########################################################################################################################################
# Create an instance on Bitcoin for Analysis
crypto_btc = Crypto("BTC-USD")
investment_decision, _ = crypto_btc.analyze_investment_opportunity(show_plot=True)
######################################################################################################################################
print(investment_decision)
wallet1.buy_crypto()
wallet1.buy_crypto()
print()
wallet1.sell_crypto()
wallet1.sell_crypto()
print()
wallet1.view_balances()
print()
wallet1.view_transaction_history()
wallet1.view_balances()
#####################################################################################################################################
print()
print(f"Total fees accumulated on the platform: {Platform.total_fees} USD")
platform.pay_taxes()
print(f"Remaining fees after taxes: {Platform.total_fees} USD")
########################################################################################################################################
# Add to Excel
save_crypto_data_to_excel(crypto_data_list)
user1.save_user_to_excel()
wallet1.save_portfolio_to_excel()
wallet2.save_portfolio_to_excel()
platform.save_accounting_to_excel()



